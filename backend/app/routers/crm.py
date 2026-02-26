"""
CRM Router — CRUD for Contacts, Companies, Pipelines, Deals, Notes, Tags.
Phase 1 Architecture: Uses Async DB, RLS injection, RBAC guards, and Audit Logging.
"""
from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete, or_
from pydantic import BaseModel
from typing import Optional, List
from uuid import UUID
import datetime
import logging

logger = logging.getLogger(__name__)

from app.db.session import get_db
from app.models.user import User
from app.models.crm import (
    CRMCompany, CRMContact, CRMPipeline, CRMDeal,
    CRMNote, CRMTag, CRMConversation
)
from app.api.deps import get_current_user
from app.core.rbac import require_permissions
from app.services.audit_service import log_audit_async

router = APIRouter()


async def _assert_owned(db: AsyncSession, model, entity_id: UUID, tenant_id: UUID, label: str):
    """
    Since RLS is enabled via session context, just filtering by ID
    is sufficient to guarantee it belongs to the tenant.
    """
    result = await db.execute(select(model).where(model.id == entity_id))
    obj = result.scalar_one_or_none()
    if not obj:
        raise HTTPException(status_code=404, detail=f"{label} not found")
    return obj


# ─── Pydantic Schemas ─────────────────────────────────────────────────

class ContactCreate(BaseModel):
    first_name: str
    last_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    company_id: Optional[UUID] = None
    position: Optional[str] = None
    linkedin_url: Optional[str] = None
    preferred_language: Optional[str] = None
    whatsapp_verified: Optional[bool] = False

class ContactUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    company_id: Optional[UUID] = None
    position: Optional[str] = None
    linkedin_url: Optional[str] = None
    preferred_language: Optional[str] = None
    whatsapp_verified: Optional[bool] = None
    payment_behavior_notes: Optional[str] = None

class CompanyCreate(BaseModel):
    name: str
    website: Optional[str] = None
    industry: Optional[str] = None
    size: Optional[str] = None
    country: Optional[str] = None
    city: Optional[str] = None
    address: Optional[str] = None
    linkedin_url: Optional[str] = None
    domain: Optional[str] = None
    risk_score: Optional[float] = None

class CompanyUpdate(BaseModel):
    name: Optional[str] = None
    website: Optional[str] = None
    industry: Optional[str] = None
    size: Optional[str] = None
    country: Optional[str] = None
    city: Optional[str] = None
    address: Optional[str] = None
    linkedin_url: Optional[str] = None
    domain: Optional[str] = None
    risk_score: Optional[float] = None

class PipelineCreate(BaseModel):
    name: str
    stages: Optional[list] = [
        {"id": "new", "name": "New"},
        {"id": "contacted", "name": "Contacted"},
        {"id": "qualified", "name": "Qualified"},
        {"id": "quoted", "name": "Quoted"},
        {"id": "negotiation", "name": "Negotiation"},
        {"id": "invoice", "name": "Invoice"},
        {"id": "paid_won", "name": "Paid/Won"}
    ]

class DealCreate(BaseModel):
    name: str
    contact_id: Optional[UUID] = None
    company_id: Optional[UUID] = None
    pipeline_id: UUID
    stage_id: str
    value: Optional[float] = 0.0
    currency: Optional[str] = "USD"
    probability: Optional[float] = 0.5
    expected_close_date: Optional[datetime.datetime] = None

class DealUpdate(BaseModel):
    name: Optional[str] = None
    stage_id: Optional[str] = None
    value: Optional[float] = None
    probability: Optional[float] = None
    status: Optional[str] = None  # open, won, lost
    expected_close_date: Optional[datetime.datetime] = None

class NoteCreate(BaseModel):
    contact_id: Optional[UUID] = None
    deal_id: Optional[UUID] = None
    company_id: Optional[UUID] = None
    content: str


# ─── Contacts ─────────────────────────────────────────────────────────

@router.get("/contacts", dependencies=[Depends(require_permissions(["crm.read"]))])
async def list_contacts(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    search: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List contacts (RLS automatically isolates to tenant)."""
    stmt = select(CRMContact)
    if search:
        search_filter = f"%{search}%"
        stmt = stmt.where(
            or_(
                CRMContact.first_name.ilike(search_filter),
                CRMContact.last_name.ilike(search_filter),
                CRMContact.email.ilike(search_filter),
                CRMContact.phone.ilike(search_filter)
            )
        )
    
    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await db.execute(count_stmt)).scalar() or 0
    
    contacts = (await db.execute(
        stmt.order_by(CRMContact.created_at.desc()).offset(skip).limit(limit)
    )).scalars().all()
    
    return {"total": total, "contacts": contacts}


@router.post("/contacts", dependencies=[Depends(require_permissions(["crm.write"]))])
async def create_contact(
    data: ContactCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if data.company_id:
        await _assert_owned(db, CRMCompany, data.company_id, current_user.current_tenant_id, "Company")

    contact = CRMContact(
        tenant_id=current_user.current_tenant_id,
        first_name=data.first_name,
        last_name=data.last_name,
        email=data.email,
        phone=data.phone,
        company_id=data.company_id,
        position=data.position,
        linkedin_url=data.linkedin_url,
    )
    db.add(contact)
    await db.commit()
    await db.refresh(contact)
    
    await log_audit_async(
        db, tenant_id=current_user.current_tenant_id, actor_user_id=current_user.id,
        action="crm_contact_created", resource_type="crm_contact", resource_id=str(contact.id)
    )
    await db.commit()
    
    return contact


@router.put("/contacts/{contact_id}", dependencies=[Depends(require_permissions(["crm.write"]))])
async def update_contact(
    contact_id: UUID,
    data: ContactUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    contact = await _assert_owned(db, CRMContact, contact_id, current_user.current_tenant_id, "Contact")

    for field, value in data.dict(exclude_unset=True).items():
        setattr(contact, field, value)

    await db.commit()
    await db.refresh(contact)
    
    await log_audit_async(
        db, tenant_id=current_user.current_tenant_id, actor_user_id=current_user.id,
        action="crm_contact_updated", resource_type="crm_contact", resource_id=str(contact.id)
    )
    await db.commit()
    
    return contact


@router.delete("/contacts/{contact_id}", dependencies=[Depends(require_permissions(["crm.admin"]))])
async def delete_contact(
    contact_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    contact = await _assert_owned(db, CRMContact, contact_id, current_user.current_tenant_id, "Contact")
    await db.delete(contact)
    
    await log_audit_async(
        db, tenant_id=current_user.current_tenant_id, actor_user_id=current_user.id,
        action="crm_contact_deleted", resource_type="crm_contact", resource_id=str(contact_id)
    )
    await db.commit()
    return {"detail": "Contact deleted"}


# ─── Companies ────────────────────────────────────────────────────────

@router.get("/companies", dependencies=[Depends(require_permissions(["crm.read"]))])
async def list_companies(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(CRMCompany)
    total = (await db.execute(select(func.count()).select_from(CRMCompany))).scalar() or 0
    companies = (await db.execute(
        stmt.order_by(CRMCompany.created_at.desc()).offset(skip).limit(limit)
    )).scalars().all()
    
    return {"total": total, "companies": companies}


@router.post("/companies", dependencies=[Depends(require_permissions(["crm.write"]))])
async def create_company(
    data: CompanyCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    company = CRMCompany(
        tenant_id=current_user.current_tenant_id,
        name=data.name,
        website=data.website,
        industry=data.industry,
        country=data.country,
        city=data.city,
        size=data.size,
        address=data.address,
        linkedin_url=data.linkedin_url,
    )
    db.add(company)
    await db.commit()
    await db.refresh(company)
    
    await log_audit_async(
        db, tenant_id=current_user.current_tenant_id, actor_user_id=current_user.id,
        action="crm_company_created", resource_type="crm_company", resource_id=str(company.id)
    )
    await db.commit()
    return company


# ─── Pipelines ────────────────────────────────────────────────────────

@router.get("/pipelines", dependencies=[Depends(require_permissions(["crm.read"]))])
async def list_pipelines(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(CRMPipeline).order_by(CRMPipeline.created_at.asc())
    pipelines = (await db.execute(stmt)).scalars().all()
    return {"pipelines": pipelines}


@router.post("/pipelines", dependencies=[Depends(require_permissions(["crm.admin"]))])
async def create_pipeline(
    data: PipelineCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    pipeline = CRMPipeline(
        tenant_id=current_user.current_tenant_id,
        name=data.name,
        stages=data.stages,
    )
    db.add(pipeline)
    await db.commit()
    await db.refresh(pipeline)
    
    await log_audit_async(
        db, tenant_id=current_user.current_tenant_id, actor_user_id=current_user.id,
        action="crm_pipeline_created", resource_type="crm_pipeline", resource_id=str(pipeline.id)
    )
    await db.commit()
    return pipeline


# ─── Deals ────────────────────────────────────────────────────────────

@router.get("/deals", dependencies=[Depends(require_permissions(["crm.read"]))])
async def list_deals(
    pipeline_id: Optional[UUID] = None,
    stage_id: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(CRMDeal)
    if pipeline_id:
        stmt = stmt.where(CRMDeal.pipeline_id == pipeline_id)
    if stage_id:
        stmt = stmt.where(CRMDeal.stage_id == stage_id)
        
    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await db.execute(count_stmt)).scalar() or 0
    
    deals = (await db.execute(
        stmt.order_by(CRMDeal.created_at.desc()).offset(skip).limit(limit)
    )).scalars().all()
    return {"total": total, "deals": deals}


@router.post("/deals", dependencies=[Depends(require_permissions(["crm.write"]))])
async def create_deal(
    data: DealCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if data.contact_id:
        await _assert_owned(db, CRMContact, data.contact_id, current_user.current_tenant_id, "Contact")
    if data.company_id:
        await _assert_owned(db, CRMCompany, data.company_id, current_user.current_tenant_id, "Company")
    await _assert_owned(db, CRMPipeline, data.pipeline_id, current_user.current_tenant_id, "Pipeline")

    deal = CRMDeal(
        tenant_id=current_user.current_tenant_id,
        name=data.name,
        contact_id=data.contact_id,
        company_id=data.company_id,
        pipeline_id=data.pipeline_id,
        stage_id=data.stage_id,
        value=data.value,
        currency=data.currency,
        probability=data.probability,
        expected_close_date=data.expected_close_date,
    )
    db.add(deal)
    await db.commit()
    await db.refresh(deal)
    
    await log_audit_async(
        db, tenant_id=current_user.current_tenant_id, actor_user_id=current_user.id,
        action="crm_deal_created", resource_type="crm_deal", resource_id=str(deal.id),
        details={"stage_id": data.stage_id, "value": data.value}
    )
    await db.commit()
    return deal


@router.put("/deals/{deal_id}", dependencies=[Depends(require_permissions(["crm.write"]))])
async def update_deal(
    deal_id: UUID,
    data: DealUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Deal stage movement and value updates."""
    deal = await _assert_owned(db, CRMDeal, deal_id, current_user.current_tenant_id, "Deal")
    
    old_stage = deal.stage_id
    updates = data.dict(exclude_unset=True)
    
    for field, value in updates.items():
        setattr(deal, field, value)
        
    if "status" in updates and updates["status"] in ["won", "lost"]:
        deal.closed_at = datetime.datetime.utcnow()

    await db.commit()
    await db.refresh(deal)
    
    audit_action = "crm_deal_stage_moved" if data.stage_id and data.stage_id != old_stage else "crm_deal_updated"
    await log_audit_async(
        db, tenant_id=current_user.current_tenant_id, actor_user_id=current_user.id,
        action=audit_action, resource_type="crm_deal", resource_id=str(deal.id),
        details={"old_stage": old_stage, "new_stage": deal.stage_id, "updates": updates}
    )
    await db.commit()
    return deal


# ─── Notes ────────────────────────────────────────────────────────────

@router.post("/notes", dependencies=[Depends(require_permissions(["crm.write"]))])
async def create_note(
    data: NoteCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if not any([data.contact_id, data.deal_id, data.company_id]):
        raise HTTPException(status_code=400, detail="Must link note to contact, deal, or company")

    note = CRMNote(
        tenant_id=current_user.current_tenant_id,
        author_id=current_user.id,
        contact_id=data.contact_id,
        deal_id=data.deal_id,
        company_id=data.company_id,
        content=data.content,
    )
    db.add(note)
    await db.commit()
    await db.refresh(note)
    
    await log_audit_async(
        db, tenant_id=current_user.current_tenant_id, actor_user_id=current_user.id,
        action="crm_note_created", resource_type="crm_note", resource_id=str(note.id)
    )
    await db.commit()
    return note

@router.get("/contacts/{contact_id}/notes", dependencies=[Depends(require_permissions(["crm.read"]))])
async def list_contact_notes(
    contact_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(CRMNote).where(
        CRMNote.contact_id == contact_id,
        CRMNote.tenant_id == current_user.current_tenant_id
    ).order_by(CRMNote.created_at.desc())
    notes = (await db.execute(stmt)).scalars().all()
    return notes

# ─── Tags ─────────────────────────────────────────────────────────────

@router.get("/tags", dependencies=[Depends(require_permissions(["crm.read"]))])
async def list_tags(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(CRMTag)
    total = (await db.execute(select(func.count()).select_from(CRMTag))).scalar() or 0
    tags = (await db.execute(stmt.offset(skip).limit(limit))).scalars().all()
    return {"total": total, "tags": tags}

class TagCreate(BaseModel):
    name: str
    color: Optional[str] = "#3B82F6"

@router.post("/tags", dependencies=[Depends(require_permissions(["crm.write"]))])
async def create_tag(
    data: TagCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tag = CRMTag(
        tenant_id=current_user.current_tenant_id,
        name=data.name,
        color=data.color,
    )
    db.add(tag)
    await db.commit()
    await db.refresh(tag)
    
    await log_audit_async(
        db, tenant_id=current_user.current_tenant_id, actor_user_id=current_user.id,
        action="crm_tag_created", resource_type="crm_tag", resource_id=str(tag.id)
    )
    await db.commit()
    return tag

# ─── Conversations / Inbox ─────────────────────────────────────────

@router.get("/conversations", dependencies=[Depends(require_permissions(["crm.read"]))])
async def list_conversations(
    contact_id: Optional[UUID] = None,
    status: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(CRMConversation)
    if contact_id:
        stmt = stmt.where(CRMConversation.contact_id == contact_id)
    if status:
        stmt = stmt.where(CRMConversation.status == status)

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await db.execute(count_stmt)).scalar() or 0
    
    conversations = (await db.execute(
        stmt.order_by(CRMConversation.last_message_at.desc()).offset(skip).limit(limit)
    )).scalars().all()
    
    return {"total": total, "conversations": conversations}

@router.get("/conversations/{conversation_id}", dependencies=[Depends(require_permissions(["crm.read"]))])
async def get_conversation(
    conversation_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    conv = await _assert_owned(db, CRMConversation, conversation_id, current_user.current_tenant_id, "Conversation")
    return conv


# ─── Stats ────────────────────────────────────────────────────────────

@router.get("/stats", dependencies=[Depends(require_permissions(["crm.read"]))])
async def crm_stats(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tid = current_user.current_tenant_id
    total_contacts = (await db.execute(select(func.count(CRMContact.id)))).scalar() or 0
    total_companies = (await db.execute(select(func.count(CRMCompany.id)))).scalar() or 0
    total_deals = (await db.execute(select(func.count(CRMDeal.id)))).scalar() or 0
    total_revenue = (await db.execute(select(func.sum(CRMDeal.value)))).scalar() or 0

    return {
        "total_contacts": total_contacts,
        "total_companies": total_companies,
        "total_deals": total_deals,
        "total_pipeline_value": float(total_revenue),
    }


# ─── Tasks ────────────────────────────────────────────────────────────

from app.models.crm import CRMTask, CRMInvoice

class TaskCreate(BaseModel):
    title: str
    description: Optional[str] = None
    company_id: Optional[UUID] = None
    contact_id: Optional[UUID] = None
    deal_id: Optional[UUID] = None
    assigned_to_id: Optional[UUID] = None
    due_date: Optional[datetime.datetime] = None
    priority: Optional[str] = "medium"

class TaskUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    due_date: Optional[datetime.datetime] = None
    priority: Optional[str] = None
    status: Optional[str] = None
    assigned_to_id: Optional[UUID] = None


@router.get("/tasks", dependencies=[Depends(require_permissions(["crm.read"]))])
async def list_tasks(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(CRMTask).where(CRMTask.tenant_id == current_user.current_tenant_id)
    if status:
        stmt = stmt.where(CRMTask.status == status)
    if priority:
        stmt = stmt.where(CRMTask.priority == priority)

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await db.execute(count_stmt)).scalar() or 0

    tasks = (await db.execute(
        stmt.order_by(CRMTask.due_date.asc().nullslast(), CRMTask.created_at.desc())
            .offset(skip).limit(limit)
    )).scalars().all()
    return {"total": total, "tasks": tasks}


@router.post("/tasks", dependencies=[Depends(require_permissions(["crm.write"]))])
async def create_task(
    data: TaskCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    task = CRMTask(
        tenant_id=current_user.current_tenant_id,
        author_id=current_user.id,
        title=data.title,
        description=data.description,
        company_id=data.company_id,
        contact_id=data.contact_id,
        deal_id=data.deal_id,
        assigned_to_id=data.assigned_to_id,
        due_date=data.due_date,
        priority=data.priority,
    )
    db.add(task)
    await db.commit()
    await db.refresh(task)

    await log_audit_async(
        db, tenant_id=current_user.current_tenant_id, actor_user_id=current_user.id,
        action="crm_task_created", resource_type="crm_task", resource_id=str(task.id)
    )
    await db.commit()
    return task


@router.put("/tasks/{task_id}", dependencies=[Depends(require_permissions(["crm.write"]))])
async def update_task(
    task_id: UUID,
    data: TaskUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    task = await _assert_owned(db, CRMTask, task_id, current_user.current_tenant_id, "Task")

    for field, value in data.dict(exclude_unset=True).items():
        setattr(task, field, value)

    if data.status == "completed":
        task.completed_at = datetime.datetime.utcnow()

    await db.commit()
    await db.refresh(task)

    await log_audit_async(
        db, tenant_id=current_user.current_tenant_id, actor_user_id=current_user.id,
        action="crm_task_updated", resource_type="crm_task", resource_id=str(task.id)
    )
    await db.commit()
    return task


# ─── Invoices (DSO / Cash Flow) ──────────────────────────────────────

class InvoiceCreate(BaseModel):
    company_id: UUID
    deal_id: Optional[UUID] = None
    invoice_number: str
    amount: float
    currency: Optional[str] = "USD"
    issued_date: datetime.datetime
    due_date: datetime.datetime

class InvoiceUpdate(BaseModel):
    status: Optional[str] = None  # open, paid, overdue, canceled
    paid_date: Optional[datetime.datetime] = None


@router.get("/invoices", dependencies=[Depends(require_permissions(["crm.read"]))])
async def list_invoices(
    status: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(CRMInvoice).where(CRMInvoice.tenant_id == current_user.current_tenant_id)
    if status:
        stmt = stmt.where(CRMInvoice.status == status)

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await db.execute(count_stmt)).scalar() or 0

    invoices = (await db.execute(
        stmt.order_by(CRMInvoice.due_date.desc()).offset(skip).limit(limit)
    )).scalars().all()

    # Calculate DSO (Days Sales Outstanding)
    open_invoices = (await db.execute(
        select(CRMInvoice).where(
            CRMInvoice.tenant_id == current_user.current_tenant_id,
            CRMInvoice.status.in_(["open", "overdue"])
        )
    )).scalars().all()

    total_outstanding = sum(float(inv.amount) for inv in open_invoices)
    dso_days = 0
    if open_invoices:
        now = datetime.datetime.utcnow()
        avg_age = sum((now - inv.issued_date).days for inv in open_invoices) / len(open_invoices)
        dso_days = round(avg_age, 1)

    return {
        "total": total,
        "invoices": invoices,
        "dso_days": dso_days,
        "total_outstanding": total_outstanding,
    }


@router.post("/invoices", dependencies=[Depends(require_permissions(["crm.write"]))])
async def create_invoice(
    data: InvoiceCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    invoice = CRMInvoice(
        tenant_id=current_user.current_tenant_id,
        company_id=data.company_id,
        deal_id=data.deal_id,
        invoice_number=data.invoice_number,
        amount=data.amount,
        currency=data.currency,
        issued_date=data.issued_date,
        due_date=data.due_date,
    )
    db.add(invoice)
    await db.commit()
    await db.refresh(invoice)

    await log_audit_async(
        db, tenant_id=current_user.current_tenant_id, actor_user_id=current_user.id,
        action="crm_invoice_created", resource_type="crm_invoice", resource_id=str(invoice.id)
    )
    await db.commit()
    return invoice


@router.put("/invoices/{invoice_id}", dependencies=[Depends(require_permissions(["crm.write"]))])
async def update_invoice(
    invoice_id: UUID,
    data: InvoiceUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    invoice = await _assert_owned(db, CRMInvoice, invoice_id, current_user.current_tenant_id, "Invoice")

    for field, value in data.dict(exclude_unset=True).items():
        setattr(invoice, field, value)

    await db.commit()
    await db.refresh(invoice)

    await log_audit_async(
        db, tenant_id=current_user.current_tenant_id, actor_user_id=current_user.id,
        action="crm_invoice_updated", resource_type="crm_invoice", resource_id=str(invoice.id)
    )
    await db.commit()
    return invoice


# ─── Bulk Import Contacts (CSV / XLSX) ───────────────────────────────

import csv
import io
import uuid as _uuid

BULK_REQUIRED_COLUMNS = {"first_name"}
BULK_ALLOWED_COLUMNS = {
    "first_name", "last_name", "email", "phone",
    "position", "linkedin_url", "preferred_language",
}

def _parse_csv_rows(raw_bytes: bytes) -> List[dict]:
    """Parse CSV bytes into list of dicts. Tries utf-8 then latin-1."""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Could not decode CSV file")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx_rows(raw_bytes: bytes) -> List[dict]:
    """Parse XLSX bytes into list of dicts using openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}" for i, h in enumerate(next(rows_iter, []))]
    results = []
    for row in rows_iter:
        d = {}
        for i, val in enumerate(row):
            if i < len(headers):
                d[headers[i]] = str(val).strip() if val is not None else ""
        results.append(d)
    wb.close()
    return results


@router.post("/contacts/bulk-import", dependencies=[Depends(require_permissions(["crm.write"]))])
async def bulk_import_contacts(
    file: UploadFile = File(...),
    event_name: Optional[str] = Form(None),
    note: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Bulk import contacts from CSV or XLSX.

    Expected columns (header row required):
      first_name (required), last_name, email, phone, position, linkedin_url, preferred_language

    Optional form fields:
      event_name — creates an 'event:<name>' tag on each imported contact
      note       — creates a CRMNote on each imported contact
    """
    fname = (file.filename or "").lower()
    ctype = (file.content_type or "").lower()
    await file.seek(0)  # ensure we read from the beginning
    raw = await file.read()
    logger.info(f"bulk-import: filename={file.filename!r}, content_type={ctype!r}, size={len(raw)}, first_bytes={raw[:20]!r}")

    is_xlsx = fname.endswith(".xlsx") or "spreadsheet" in ctype or "xlsx" in ctype
    is_csv = fname.endswith(".csv") or "csv" in ctype or "text/" in ctype

    if is_xlsx:
        try:
            rows = _parse_xlsx_rows(raw)
        except Exception as e:
            logger.exception(f"bulk-import XLSX parse error (size={len(raw)}): {e}")
            raise HTTPException(status_code=400, detail=f"Failed to parse XLSX: {e}")
    elif is_csv:
        try:
            rows = _parse_csv_rows(raw)
        except Exception as e:
            logger.exception(f"bulk-import CSV parse error (size={len(raw)}): {e}")
            raise HTTPException(status_code=400, detail=f"Failed to parse CSV: {e}")
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported file type '{fname}'. Upload .csv or .xlsx")

    if not rows:
        raise HTTPException(status_code=400, detail="File contains no data rows")

    # Normalize column names
    normalized = []
    for row in rows:
        norm = {k.strip().lower().replace(" ", "_"): (v.strip() if isinstance(v, str) else v) for k, v in row.items()}
        normalized.append(norm)

    # Validate required columns exist
    sample_keys = set(normalized[0].keys())
    missing = BULK_REQUIRED_COLUMNS - sample_keys
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

    tenant_id = current_user.current_tenant_id
    created = 0
    skipped = 0
    errors = []

    for idx, row in enumerate(normalized, start=2):  # row 2 = first data row (1=header)
        first_name = (row.get("first_name") or "").strip()
        if not first_name:
            skipped += 1
            errors.append({"row": idx, "error": "Missing first_name"})
            continue

        # Build contact
        contact = CRMContact(
            tenant_id=tenant_id,
            first_name=first_name,
            last_name=(row.get("last_name") or "").strip() or None,
            email=(row.get("email") or "").strip() or None,
            phone=(row.get("phone") or "").strip() or None,
            position=(row.get("position") or "").strip() or None,
            linkedin_url=(row.get("linkedin_url") or "").strip() or None,
            preferred_language=(row.get("preferred_language") or "").strip() or None,
        )
        db.add(contact)
        await db.flush()  # get contact.id

        # Event tag
        if event_name and event_name.strip():
            tag = CRMTag(
                tenant_id=tenant_id,
                name=f"event:{event_name.strip()}",
                color="#D4AF37",
                entity_type="contact",
                entity_id=contact.id,
            )
            db.add(tag)

        # Note
        if note and note.strip():
            n = CRMNote(
                tenant_id=tenant_id,
                author_id=current_user.id,
                contact_id=contact.id,
                content=note.strip(),
            )
            db.add(n)

        created += 1

    await db.commit()

    await log_audit_async(
        db, tenant_id=tenant_id, actor_user_id=current_user.id,
        action="crm_contacts_bulk_imported", resource_type="crm_contact",
        resource_id="bulk",
        details={"created": created, "skipped": skipped, "event_name": event_name}
    )
    await db.commit()

    return {
        "created": created,
        "skipped": skipped,
        "total_rows": len(normalized),
        "errors": errors[:20],  # limit error detail
    }
